from flask import Blueprint, request, jsonify, g
import traceback
from io import BytesIO

from database.database import conectar
from utils.auth_middleware import auth_required

excel_bp = Blueprint("excel", __name__, url_prefix="/excel")

# Colunas aceitas por campo (variações de nome)
MAPA_COLUNAS = {
    "produto":    ["produto", "produtos", "product", "nome", "name", "descricao", "descrição", "item"],
    "quantidade": ["quantidade", "qtd", "qty", "quantity", "estoque", "qtde", "saldo",
                   "entradas", "saidas", "saídas"],
    "valor":      ["valor", "value", "preco", "preço", "price", "custo", "cost", "vl", "vr",
                   "media de custo", "média de custo", "custo unitario", "custo unitário",
                   "preco unitario", "preço unitário"],
    "fornecedor": ["fornecedor", "supplier", "vendor", "fabricante"],
    "contato":    ["contato", "contact", "telefone", "tel", "fone", "phone", "e-mail", "email"],
}


def normalizar(texto):
    import unicodedata
    texto = str(texto).strip().lower()
    return unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("ascii")


def mapear_cabecalho(cabecalho):
    """Retorna dict {campo_interno: indice_coluna}."""
    mapa = {}
    for idx, col in enumerate(cabecalho):
        col_norm = normalizar(col)
        for campo, variantes in MAPA_COLUNAS.items():
            if col_norm in [normalizar(v) for v in variantes]:
                if campo not in mapa:
                    mapa[campo] = idx
    return mapa


def detectar_cabecalho(linhas, max_busca=10):
    """
    Percorre as primeiras linhas procurando aquela que parece um cabeçalho
    (tem mais células não-nulas e pelo menos uma coluna mapeável).
    Retorna (indice_linha, mapa_colunas) ou (None, {}).
    """
    for i, linha in enumerate(linhas[:max_busca]):
        celulas = [str(c).strip() if c is not None else "" for c in linha]
        nao_vazias = [c for c in celulas if c]
        if len(nao_vazias) < 2:
            continue
        mapa = mapear_cabecalho(celulas)
        if "produto" in mapa or "quantidade" in mapa:
            return i, mapa
    return None, {}


def linhas_para_preview(linhas, max_linhas=5):
    """
    Converte linhas do openpyxl para strings serializáveis,
    filtrando colunas totalmente vazias.
    """
    import datetime

    def cel_str(v):
        if v is None:
            return ""
        if isinstance(v, datetime.datetime):
            return v.strftime("%d/%m/%Y")
        return str(v).strip()

    todas = [tuple(cel_str(c) for c in linha) for linha in linhas[:max_linhas + 2]]
    n_cols = max((len(r) for r in todas), default=0)

    # Só mantém colunas que têm algum valor
    colunas_com_dado = [
        j for j in range(n_cols)
        if any(j < len(r) and r[j] for r in todas)
    ]

    resultado = []
    for row in todas[:max_linhas]:
        resultado.append([row[j] if j < len(row) else "" for j in colunas_com_dado])

    return resultado


# =========================
# ROTA: PREVIEW DAS ABAS
# =========================
@excel_bp.route("/preview", methods=["POST"])
@auth_required
def preview_excel():
    """
    Recebe o arquivo e retorna a lista de abas com preview das primeiras linhas,
    quais campos foram mapeados e se a aba parece importável.
    """
    try:
        arquivo = request.files.get("arquivo")
        if not arquivo:
            return jsonify({"erro": "Arquivo não enviado"}), 400

        nome_arquivo = arquivo.filename or ""
        extensao = nome_arquivo.rsplit(".", 1)[-1].lower() if "." in nome_arquivo else ""

        if extensao not in ("xlsx", "xls", "ods"):
            return jsonify({"erro": "Formato inválido. Envie .xlsx, .xls ou .ods"}), 400

        conteudo = arquivo.read()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
        except Exception as e:
            return jsonify({"erro": f"Erro ao ler planilha: {str(e)}"}), 400

        abas = []
        for nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            linhas = []
            for row in ws.iter_rows(values_only=True):
                linhas.append(row)
                if len(linhas) >= 15:
                    break

            idx_cabecalho, mapa = detectar_cabecalho(linhas)
            tem_produto = "produto" in mapa

            # Preview a partir do cabeçalho detectado
            inicio = idx_cabecalho if idx_cabecalho is not None else 0
            preview = linhas_para_preview(linhas[inicio:], max_linhas=5)

            abas.append({
                "nome":            nome_aba,
                "tem_produto":     tem_produto,
                "cabecalho_linha": idx_cabecalho,
                "colunas_mapeadas": list(mapa.keys()),
                "preview":         preview,
            })

        wb.close()

        return jsonify({"abas": abas}), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500


# =========================
# ROTA: IMPORTAR ABA ESCOLHIDA
# =========================
@excel_bp.route("/importar", methods=["POST"])
@auth_required
def importar_excel():
    try:
        tenant_id = g.usuario["tenant_id"]

        arquivo = request.files.get("arquivo")
        if not arquivo:
            return jsonify({"erro": "Arquivo não enviado"}), 400

        nome_arquivo = arquivo.filename or ""
        extensao = nome_arquivo.rsplit(".", 1)[-1].lower() if "." in nome_arquivo else ""

        if extensao not in ("xlsx", "xls", "ods"):
            return jsonify({"erro": "Formato inválido. Envie .xlsx, .xls ou .ods"}), 400

        # Aba escolhida pelo usuário via form field
        aba_escolhida = request.form.get("aba", "").strip()

        conteudo = arquivo.read()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)

            if aba_escolhida and aba_escolhida in wb.sheetnames:
                ws = wb[aba_escolhida]
            else:
                ws = wb.active

            linhas = list(ws.iter_rows(values_only=True))
            wb.close()

        except Exception as e:
            return jsonify({"erro": f"Erro ao ler planilha: {str(e)}"}), 400

        if not linhas or len(linhas) < 2:
            return jsonify({"erro": "Planilha vazia ou sem dados além do cabeçalho"}), 400

        # Detecta onde está o cabeçalho automaticamente
        idx_cabecalho, mapa = detectar_cabecalho(linhas)

        if idx_cabecalho is None or "produto" not in mapa:
            cabecalho_raw = [str(c) if c is not None else "" for c in (linhas[0] if linhas else [])]
            return jsonify({
                "erro": "Não foi possível encontrar a coluna 'produto' nessa aba. Verifique o cabeçalho.",
                "cabecalho_encontrado": cabecalho_raw
            }), 400

        dados_linhas = linhas[idx_cabecalho + 1:]

        if not dados_linhas:
            return jsonify({"erro": "Aba sem dados após o cabeçalho"}), 400

        conn = conectar()
        cursor = conn.cursor()

        importados = []
        ignorados  = []

        import datetime

        for i, linha in enumerate(dados_linhas, start=idx_cabecalho + 2):

            def cel(campo, padrao="", _linha=linha):
                idx = mapa.get(campo)
                if idx is None or idx >= len(_linha):
                    return padrao
                val = _linha[idx]
                return val if val is not None else padrao

            # Nome do produto
            nome_raw = cel("produto", "")
            if isinstance(nome_raw, datetime.datetime):
                ignorados.append(f"Linha {i}: valor é uma data, ignorado")
                continue
            nome = str(nome_raw).strip()
            if not nome or nome.lower() in ("none", "nan", ""):
                ignorados.append(f"Linha {i}: produto vazio")
                continue

            # Quantidade
            try:
                quantidade = float(str(cel("quantidade", 0)).replace(",", ".") or 0)
            except (ValueError, TypeError):
                quantidade = 0.0

            # Valor
            try:
                valor = float(
                    str(cel("valor", 0)).replace(",", ".").replace("R$", "").strip() or 0
                )
            except (ValueError, TypeError):
                valor = 0.0

            fornecedor = str(cel("fornecedor", "")).strip()
            contato    = str(cel("contato",    "")).strip()

            # Verifica se produto já existe (preço médio ponderado)
            cursor.execute(
                "SELECT id, quantidade, valor FROM produtos WHERE produto = %s AND tenant_id = %s",
                (nome, tenant_id)
            )
            existente = cursor.fetchone()

            if existente:
                qtd_atual   = float(existente["quantidade"] or 0)
                valor_atual = float(existente["valor"] or 0)
                nova_qtd    = qtd_atual + quantidade

                if nova_qtd > 0 and quantidade > 0:
                    novo_valor = ((qtd_atual * valor_atual) + (quantidade * valor)) / nova_qtd
                else:
                    novo_valor = valor_atual

                cursor.execute("""
                    UPDATE produtos
                    SET quantidade = %s,
                        valor      = %s,
                        fornecedor = CASE WHEN %s != '' THEN %s ELSE fornecedor END,
                        contato    = CASE WHEN %s != '' THEN %s ELSE contato END
                    WHERE produto = %s AND tenant_id = %s
                """, (
                    nova_qtd, round(novo_valor, 4),
                    fornecedor, fornecedor,
                    contato,    contato,
                    nome, tenant_id
                ))
            else:
                cursor.execute("""
                    INSERT INTO produtos (tenant_id, produto, quantidade, valor, fornecedor, contato)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (tenant_id, nome, quantidade, valor, fornecedor, contato))

            importados.append(nome)

        conn.commit()
        conn.close()

        return jsonify({
            "msg":              "Planilha importada com sucesso",
            "aba":              aba_escolhida or "ativa",
            "total_importados": len(importados),
            "produtos":         importados,
            "ignorados":        ignorados
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500